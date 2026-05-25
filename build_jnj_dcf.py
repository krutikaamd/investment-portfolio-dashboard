"""
JNJ DCF Valuation Model — Full Workings
Generates: JNJ DCF Valuation Model.xlsx
Mirrors the AMD DCF template structure: Cover / Inputs / Model / Outputs
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
import math

wb = Workbook()

# ── Colour palette ─────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"
C_MID_BLUE    = "2E75B6"
C_LIGHT_BLUE  = "BDD7EE"
C_GREEN       = "375623"
C_LIGHT_GREEN = "E2EFDA"
C_ORANGE      = "C55A11"
C_YELLOW      = "FFF2CC"
C_RED_LIGHT   = "FCE4D6"
C_GREY        = "F2F2F2"
C_WHITE       = "FFFFFF"
C_BLACK       = "000000"

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def font(bold=False, colour=C_BLACK, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=colour, size=size, italic=italic)

def border_thin():
    s = Side(style="thin", color=C_BLACK)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    s = Side(style="thin", color="BFBFBF")
    return Border(bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def header_row(ws, row, cells, bg=C_DARK_BLUE, fg=C_WHITE):
    for col, val in cells:
        c = ws.cell(row=row, column=col, value=val)
        c.font = font(bold=True, colour=fg, size=10)
        c.fill = fill(bg)
        c.alignment = align("center")
        c.border = border_thin()

def section_title(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=True, colour=C_WHITE, size=11)
    c.fill = fill(C_MID_BLUE)
    c.alignment = align("left")

def label(ws, row, col, val, bold=False, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, italic=italic)
    c.alignment = align("left")
    return c

def num(ws, row, col, val, fmt=None, bg=None, bold=False, colour=C_BLACK):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, colour=colour)
    c.alignment = align("right")
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = fill(bg)
    return c

def pct_fmt():  return '0.0%'
def pct1_fmt(): return '0.0%'
def pct2_fmt(): return '0.00%'
def usd_fmt():  return '$#,##0.0'
def usd0_fmt(): return '$#,##0'
def usd2_fmt(): return '$#,##0.00'
def dec_fmt():  return '#,##0.0'
def dec2_fmt(): return '#,##0.00'

# ═══════════════════════════════════════════════════════════════════════════════
# 1. COVER SHEET
# ═══════════════════════════════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "Cover"
ws_cover.sheet_view.showGridLines = False
ws_cover.column_dimensions["A"].width = 3
ws_cover.column_dimensions["B"].width = 45
ws_cover.column_dimensions["C"].width = 28
ws_cover.column_dimensions["D"].width = 20

# Title
ws_cover.row_dimensions[5].height = 36
c = ws_cover.cell(row=5, column=2, value="JNJ DCF VALUATION MODEL")
c.font = Font(name="Calibri", bold=True, size=20, color=C_WHITE)
c.fill = fill(C_DARK_BLUE)
c.alignment = align("left", "center")
ws_cover.merge_cells("B5:D5")

ws_cover.row_dimensions[6].height = 18
c = ws_cover.cell(row=6, column=2,
    value="Johnson & Johnson (NYSE: JNJ)  |  10-Year DCF Analysis  |  May 21, 2026")
c.font = font(italic=True, colour=C_WHITE, size=11)
c.fill = fill(C_MID_BLUE)
c.alignment = align("left", "center")
ws_cover.merge_cells("B6:D6")

# Table of contents
for r, txt in enumerate([
    ("",),
    ("TABLE OF CONTENTS",),
    ("",),
    ("Inputs",  "Assumptions, WACC construction, revenue drivers"),
    ("Model",   "Year-by-year FCF projections — all three scenarios"),
    ("Outputs", "DCF outputs, sensitivity table, verdict"),
], start=8):
    if len(txt) == 1:
        c = ws_cover.cell(row=r, column=2, value=txt[0])
        if txt[0] == "TABLE OF CONTENTS":
            c.font = font(bold=True, size=12, colour=C_MID_BLUE)
    else:
        c = ws_cover.cell(row=r, column=2, value=txt[0])
        c.font = font(bold=True)
        c2 = ws_cover.cell(row=r, column=3, value=txt[1])
        c2.font = font(italic=True, colour="595959")

# Summary box
for r, (k, v) in enumerate([
    ("Valuation Date",    "May 21, 2026"),
    ("Company",           "Johnson & Johnson"),
    ("Ticker",            "NYSE: JNJ"),
    ("Current Price",     "$229.32"),
    ("Market Cap",        "$553B"),
    ("Shares (diluted)",  "2.41B"),
    ("Net Debt",          "$32.9B"),
    ("Beta (5-yr)",       "0.33"),
    ("Analyst Consensus", "$252 / Moderate Buy"),
    ("Dividend Yield",    "3.05%"),
    ("Base DCF Value",    "$217 per share"),
    ("VERDICT",           "HOLD — Do Not Add at $229"),
], start=8):
    ck = ws_cover.cell(row=r, column=5, value=k)
    ck.font = font(bold=True, colour=C_WHITE)
    ck.fill = fill(C_MID_BLUE)
    ck.alignment = align("left")
    cv = ws_cover.cell(row=r, column=6, value=v)
    cv.font = font(colour=C_BLACK)
    cv.fill = fill(C_LIGHT_BLUE)
    cv.alignment = align("left")
    if k == "VERDICT":
        ck.fill = fill(C_ORANGE)
        cv.fill = fill(C_YELLOW)
        cv.font = font(bold=True)
ws_cover.column_dimensions["E"].width = 22
ws_cover.column_dimensions["F"].width = 30

ws_cover.cell(row=22, column=2,
    value="Disclaimer: This model is for informational purposes only. Not investment advice.")
ws_cover.cell(row=22, column=2).font = font(italic=True, colour="595959", size=9)

# ═══════════════════════════════════════════════════════════════════════════════
# 2. INPUTS SHEET
# ═══════════════════════════════════════════════════════════════════════════════
ws_in = wb.create_sheet("Inputs")
ws_in.sheet_view.showGridLines = False
ws_in.column_dimensions["A"].width = 3
ws_in.column_dimensions["B"].width = 42
ws_in.column_dimensions["C"].width = 16
ws_in.column_dimensions["D"].width = 16
ws_in.column_dimensions["E"].width = 16
ws_in.column_dimensions["F"].width = 16
ws_in.column_dimensions["G"].width = 16
ws_in.column_dimensions["H"].width = 16
ws_in.column_dimensions["I"].width = 16
ws_in.column_dimensions["J"].width = 16
ws_in.column_dimensions["K"].width = 16
ws_in.column_dimensions["L"].width = 16

r = 1

# ── Section: Company & Market Data ────────────────────────────────────────────
section_title(ws_in, r, 2, " COMPANY & MARKET DATA")
ws_in.merge_cells(f"B{r}:H{r}")
r += 1

for k, v, fmt in [
    ("Current Stock Price (USD)",        229.32,  usd2_fmt()),
    ("Analyst Consensus Price Target",   252.00,  usd2_fmt()),
    ("Shares Outstanding (billions)",    2.41,    dec2_fmt()),
    ("Market Capitalisation (USD B)",    553.0,   usd_fmt()),
    ("Beta (5-yr monthly, Yahoo)",       0.33,    dec2_fmt()),
    ("Dividend Yield (%)",               0.0305,  pct1_fmt()),
    ("Consecutive Dividend Increases",   63,      "0"),
]:
    label(ws_in, r, 2, k)
    num(ws_in, r, 3, v, fmt=fmt, bg=C_LIGHT_BLUE)
    r += 1

r += 1

# ── Section: Balance Sheet ────────────────────────────────────────────────────
section_title(ws_in, r, 2, " BALANCE SHEET INPUTS (USD billions)")
ws_in.merge_cells(f"B{r}:H{r}")
r += 1

for k, v, fmt in [
    ("Cash & Short-Term Investments",    22.1,  usd_fmt()),
    ("Total Debt",                       55.0,  usd_fmt()),
    ("Net Debt (Debt − Cash)",           32.9,  usd_fmt()),
    ("Total Equity (Market Cap)",       553.0,  usd_fmt()),
    ("Equity Weight (in capital structure)", 553/(553+55), pct1_fmt()),
    ("Debt Weight (in capital structure)",    55/(553+55),  pct1_fmt()),
]:
    label(ws_in, r, 2, k)
    num(ws_in, r, 3, v, fmt=fmt, bg=C_LIGHT_BLUE)
    r += 1

r += 1

# ── Section: WACC Construction ───────────────────────────────────────────────
section_title(ws_in, r, 2, " WACC CONSTRUCTION")
ws_in.merge_cells(f"B{r}:H{r}")
r += 1

header_row(ws_in, r, [(2,"Component"),(3,"Value"),(4,"Notes")], bg=C_MID_BLUE)
r += 1

wacc_rows = [
    ("Risk-Free Rate (10-yr US Treasury)",    0.0429, "May 2026"),
    ("Equity Risk Premium (US market)",       0.0490, "Damodaran ERP estimate"),
    ("Country Risk Premium (USA)",            0.0050, "USA-specific premium"),
    ("Beta (5-year monthly)",                 0.33,   "Yahoo Finance"),
    ("CAPM Cost of Equity",                   0.0429+0.33*0.0490+0.0050, "Rf + β×ERP + CRP"),
    ("Healthcare Sector WACC Floor Applied",  0.075,  "Adjusted for litigation/patent risk"),
    ("Pre-Tax Cost of Debt",                  0.0450, "A+ credit rating"),
    ("Effective Tax Rate",                    0.170,  "FY2025 effective rate"),
    ("After-Tax Cost of Debt",                0.0450*(1-0.170), "Pre-tax × (1 − tax)"),
    ("WACC — Bull Case",                      0.075,  "Low beta premium; see Bull scenario"),
    ("WACC — Base Case",                      0.080,  "Standard healthcare WACC"),
    ("WACC — Bear Case",                      0.095,  "Elevated risk premium"),
]
for k, v, note in wacc_rows:
    label(ws_in, r, 2, k)
    num(ws_in, r, 3, v, fmt=pct2_fmt() if isinstance(v, float) else "0", bg=C_LIGHT_BLUE)
    ws_in.cell(row=r, column=4, value=note).font = font(italic=True, colour="595959")
    r += 1

r += 1

# ── Section: Revenue & FCF Drivers ───────────────────────────────────────────
section_title(ws_in, r, 2, " REVENUE & FCF DRIVER ASSUMPTIONS")
ws_in.merge_cells(f"B{r}:L{r}")
r += 1

years = list(range(2026, 2036))
year_labels = [str(y)+"E" for y in years]
header_row(ws_in, r,
    [(2,"Driver"),(3,"Scenario")] + [(i+4, y) for i, y in enumerate(year_labels)],
    bg=C_MID_BLUE)
r += 1

bull_rev_growth  = [0.069, 0.063, 0.084, 0.069, 0.081, 0.067, 0.056, 0.046, 0.032, 0.025]
base_rev_growth  = [0.069, 0.056, 0.081, 0.056, 0.082, 0.052, 0.043, 0.040, 0.030, 0.027]
bear_rev_growth  = [0.061, 0.045, 0.040, 0.040, 0.035, 0.030, 0.030, 0.025, 0.025, 0.020]
bull_fcf_margins = [0.220, 0.230, 0.240, 0.250, 0.260, 0.270, 0.270, 0.280, 0.280, 0.290]
base_fcf_margins = [0.210, 0.215, 0.220, 0.220, 0.225, 0.230, 0.230, 0.235, 0.235, 0.240]
bear_fcf_margins = [0.195, 0.200, 0.200, 0.200, 0.205, 0.205, 0.210, 0.210, 0.210, 0.210]

for label_txt, scenario, data, bg in [
    ("Revenue Growth Rate", "Bull",  bull_rev_growth,  C_LIGHT_GREEN),
    ("Revenue Growth Rate", "Base",  base_rev_growth,  C_GREY),
    ("Revenue Growth Rate", "Bear",  bear_rev_growth,  C_RED_LIGHT),
    ("FCF Margin",          "Bull",  bull_fcf_margins, C_LIGHT_GREEN),
    ("FCF Margin",          "Base",  base_fcf_margins, C_GREY),
    ("FCF Margin",          "Bear",  bear_fcf_margins, C_RED_LIGHT),
]:
    label(ws_in, r, 2, label_txt)
    label(ws_in, r, 3, scenario, bold=True)
    for i, v in enumerate(data):
        num(ws_in, r, i+4, v, fmt=pct1_fmt(), bg=bg)
    r += 1

r += 1

# ── Section: Terminal Value Assumptions ───────────────────────────────────────
section_title(ws_in, r, 2, " TERMINAL VALUE ASSUMPTIONS")
ws_in.merge_cells(f"B{r}:H{r}")
r += 1

for k, v, fmt, note in [
    ("Terminal Growth Rate — Bull",  0.030, pct1_fmt(), "Long-run GDP proxy"),
    ("Terminal Growth Rate — Base",  0.030, pct1_fmt(), "Conservative; below WACC by 500bps"),
    ("Terminal Growth Rate — Bear",  0.020, pct1_fmt(), "Slow maturation"),
    ("Terminal Year FCF — Bull",     48.4,  usd_fmt(),  "USD billions; 2035E"),
    ("Terminal Year FCF — Base",     38.0,  usd_fmt(),  "USD billions; 2035E"),
    ("Terminal Year FCF — Bear",     27.9,  usd_fmt(),  "USD billions; 2035E"),
]:
    label(ws_in, r, 2, k)
    num(ws_in, r, 3, v, fmt=fmt, bg=C_LIGHT_BLUE)
    ws_in.cell(row=r, column=4, value=note).font = font(italic=True, colour="595959")
    r += 1

# ═══════════════════════════════════════════════════════════════════════════════
# 3. MODEL SHEET
# ═══════════════════════════════════════════════════════════════════════════════
ws_mod = wb.create_sheet("Model")
ws_mod.sheet_view.showGridLines = False
ws_mod.column_dimensions["A"].width = 3
ws_mod.column_dimensions["B"].width = 38
ws_mod.column_dimensions["C"].width = 14

for col in range(4, 16):
    ws_mod.column_dimensions[get_column_letter(col)].width = 13

# Historical and projection data
hist_yrs = ["2023A", "2024A", "2025A"]
proj_yrs = [str(y)+"E" for y in range(2026, 2036)]
all_yrs  = hist_yrs + proj_yrs

hist_rev = [85.2, 88.8, 94.2]
hist_fcf = [18.6, 19.8, 19.7]
hist_fcfm = [0.218, 0.223, 0.209]
hist_opM  = [0.250, 0.263, 0.270]

bull_rev  = [100.7, 107.0, 116.0, 124.0, 134.0, 143.0, 151.0, 158.0, 163.0, 167.0]
base_rev  = [100.7, 106.3, 114.9, 121.3, 131.2, 138.0, 143.9, 149.7, 154.2, 158.3]
bear_rev  = [ 99.9, 104.4, 108.6, 112.9, 116.9, 120.4, 124.0, 127.1, 130.3, 132.9]

bull_fcf_val = [22.2, 24.6, 27.8, 31.0, 34.8, 38.6, 40.8, 44.2, 45.6, 48.4]
base_fcf_val = [21.1, 22.9, 25.3, 26.7, 29.5, 31.7, 33.1, 35.2, 36.2, 38.0]
bear_fcf_val = [19.5, 20.9, 21.7, 22.6, 24.0, 24.7, 26.0, 26.7, 27.4, 27.9]

WACC_BULL = 0.075
WACC_BASE = 0.080
WACC_BEAR = 0.095
G_BULL    = 0.030
G_BASE    = 0.030
G_BEAR    = 0.020
NET_DEBT  = 32.9
SHARES    = 2.41

def pv_fcfs(fcf_list, wacc):
    return sum(f / (1+wacc)**(i+1) for i, f in enumerate(fcf_list))

def terminal_value(fcf_last, wacc, g):
    return fcf_last * (1+g) / (wacc - g)

def pv_tv(tv, wacc, n=10):
    return tv / (1+wacc)**n

bull_pv_fcf = pv_fcfs(bull_fcf_val, WACC_BULL)
base_pv_fcf = pv_fcfs(base_fcf_val, WACC_BASE)
bear_pv_fcf = pv_fcfs(bear_fcf_val, WACC_BEAR)

bull_tv = terminal_value(bull_fcf_val[-1], WACC_BULL, G_BULL)
base_tv = terminal_value(base_fcf_val[-1], WACC_BASE, G_BASE)
bear_tv = terminal_value(bear_fcf_val[-1], WACC_BEAR, G_BEAR)

bull_pv_tv  = pv_tv(bull_tv, WACC_BULL)
base_pv_tv  = pv_tv(base_tv, WACC_BASE)
bear_pv_tv  = pv_tv(bear_tv, WACC_BEAR)

bull_ev     = bull_pv_fcf + bull_pv_tv
base_ev     = base_pv_fcf + base_pv_tv
bear_ev     = bear_pv_fcf + bear_pv_tv

bull_eq     = bull_ev - NET_DEBT
base_eq     = base_ev - NET_DEBT
bear_eq     = bear_ev - NET_DEBT

bull_ps     = bull_eq / SHARES
base_ps     = base_eq / SHARES
bear_ps     = bear_eq / SHARES

r = 1

# ── Historical Income Statement ───────────────────────────────────────────────
section_title(ws_mod, r, 2, " HISTORICAL INCOME STATEMENT (USD billions)")
ws_mod.merge_cells(f"B{r}:F{r}")
r += 1

header_row(ws_mod, r,
    [(2,"Metric")] + [(i+3, y) for i, y in enumerate(hist_yrs)],
    bg=C_DARK_BLUE)
r += 1

hist_rows = [
    ("Revenue (USD B)",          hist_rev,  usd_fmt()),
    ("YoY Revenue Growth",        [None, 0.043, 0.062], pct1_fmt()),
    ("Operating Margin",          hist_opM,  pct1_fmt()),
    ("Free Cash Flow (USD B)",    hist_fcf,  usd_fmt()),
    ("FCF Margin",                hist_fcfm, pct1_fmt()),
    ("Adj. EPS (USD)",           [9.19, 9.98, 10.79], usd2_fmt()),
    ("Dividend Per Share (USD)",  [4.70, 4.96, 5.12],  usd2_fmt()),
]
for label_txt, vals, fmt in hist_rows:
    label(ws_mod, r, 2, label_txt)
    for i, v in enumerate(vals):
        if v is None:
            ws_mod.cell(row=r, column=i+3, value="—")
        else:
            num(ws_mod, r, i+3, v, fmt=fmt, bg=C_GREY)
    r += 1

r += 1

# ── Three Scenario Projections ────────────────────────────────────────────────
for scenario, s_rev, s_fcf, s_revg, s_fcfm, wacc, g, pv_f, tv, pv_t, ev, eq, ps, bg, bg2 in [
    ("BULL CASE  |  WACC 7.5%  |  Terminal g 3.0%  |  FCF Margin 22%→29%",
     bull_rev, bull_fcf_val, bull_rev_growth, bull_fcf_margins,
     WACC_BULL, G_BULL, bull_pv_fcf, bull_tv, bull_pv_tv, bull_ev, bull_eq, bull_ps,
     C_GREEN, C_LIGHT_GREEN),
    ("BASE CASE  |  WACC 8.0%  |  Terminal g 3.0%  |  FCF Margin 21%→24%",
     base_rev, base_fcf_val, base_rev_growth, base_fcf_margins,
     WACC_BASE, G_BASE, base_pv_fcf, base_tv, base_pv_tv, base_ev, base_eq, base_ps,
     C_MID_BLUE, C_LIGHT_BLUE),
    ("BEAR CASE  |  WACC 9.5%  |  Terminal g 2.0%  |  FCF Margin 19%→21%",
     bear_rev, bear_fcf_val, bear_rev_growth, bear_fcf_margins,
     WACC_BEAR, G_BEAR, bear_pv_fcf, bear_tv, bear_pv_tv, bear_ev, bear_eq, bear_ps,
     C_ORANGE, C_RED_LIGHT),
]:
    # Section header
    c = ws_mod.cell(row=r, column=2, value=f" {scenario}")
    c.font = font(bold=True, colour=C_WHITE, size=10)
    c.fill = fill(bg)
    c.alignment = align("left")
    ws_mod.merge_cells(f"B{r}:{get_column_letter(3+len(proj_yrs)-1)}{r}")
    r += 1

    # Year headers
    header_row(ws_mod, r,
        [(2,"Metric")] + [(i+3, y) for i, y in enumerate(proj_yrs)],
        bg=C_DARK_BLUE)
    r += 1

    # Data rows
    pv_row = [v / (1+wacc)**(i+1) for i, v in enumerate(s_fcf)]
    disc_factors = [1/(1+wacc)**(i+1) for i in range(10)]

    scenario_rows = [
        ("Revenue (USD B)",          s_rev,       usd_fmt()),
        ("YoY Revenue Growth",       s_revg,      pct1_fmt()),
        ("FCF Margin",               s_fcfm,      pct1_fmt()),
        ("Free Cash Flow (USD B)",   s_fcf,       usd_fmt()),
        ("Discount Factor",          disc_factors, "0.0000"),
        ("PV of FCF (USD B)",        pv_row,       usd_fmt()),
    ]
    for label_txt, vals, fmt in scenario_rows:
        label(ws_mod, r, 2, label_txt)
        for i, v in enumerate(vals):
            num(ws_mod, r, i+3, v, fmt=fmt, bg=bg2)
        r += 1

    # Terminal value block
    r += 1
    label(ws_mod, r, 2, "Terminal Year FCF (USD B)",     bold=True)
    num(ws_mod, r, 3, s_fcf[-1], fmt=usd_fmt(), bg=bg2, bold=True)
    r += 1
    label(ws_mod, r, 2, "Terminal Growth Rate (g)",       bold=True)
    num(ws_mod, r, 3, g, fmt=pct1_fmt(), bg=bg2, bold=True)
    r += 1
    label(ws_mod, r, 2, "WACC",                           bold=True)
    num(ws_mod, r, 3, wacc, fmt=pct1_fmt(), bg=bg2, bold=True)
    r += 1
    label(ws_mod, r, 2, "Terminal Value  [FCF×(1+g)/(WACC−g)]  (USD B)", bold=True)
    num(ws_mod, r, 3, tv, fmt=usd_fmt(), bg=bg2, bold=True)
    r += 1
    label(ws_mod, r, 2, f"PV of Terminal Value  [TV/(1+WACC)^10]  (USD B)", bold=True)
    num(ws_mod, r, 3, pv_t, fmt=usd_fmt(), bg=bg2, bold=True)
    r += 1
    r += 1
    label(ws_mod, r, 2, "Sum of PV of FCFs  (USD B)",    bold=True)
    num(ws_mod, r, 3, pv_f, fmt=usd_fmt(), bg=bg2, bold=True)
    r += 1
    label(ws_mod, r, 2, "PV of Terminal Value  (USD B)", bold=True)
    num(ws_mod, r, 3, pv_t, fmt=usd_fmt(), bg=bg2, bold=True)
    r += 1
    label(ws_mod, r, 2, "Enterprise Value  (USD B)",     bold=True)
    num(ws_mod, r, 3, ev,  fmt=usd_fmt(), bg=bg2, bold=True)
    r += 1
    label(ws_mod, r, 2, "(−) Net Debt  (USD B)",        bold=True)
    num(ws_mod, r, 3, NET_DEBT, fmt=usd_fmt(), bg=bg2, bold=True)
    r += 1
    label(ws_mod, r, 2, "Equity Value  (USD B)",         bold=True)
    num(ws_mod, r, 3, eq,  fmt=usd_fmt(), bg=bg2, bold=True)
    r += 1
    label(ws_mod, r, 2, "Shares Outstanding  (B)",       bold=True)
    num(ws_mod, r, 3, SHARES, fmt=dec2_fmt(), bg=bg2, bold=True)
    r += 1

    c = ws_mod.cell(row=r, column=2, value="EQUITY VALUE PER SHARE  (USD)")
    c.font = font(bold=True, colour=C_WHITE, size=12)
    c.fill = fill(bg)
    c.alignment = align("left")
    c2 = ws_mod.cell(row=r, column=3, value=round(ps, 2))
    c2.font = font(bold=True, colour=C_WHITE, size=12)
    c2.fill = fill(bg)
    c2.number_format = usd2_fmt()
    c2.alignment = align("right")
    r += 2

# ═══════════════════════════════════════════════════════════════════════════════
# 4. OUTPUTS SHEET
# ═══════════════════════════════════════════════════════════════════════════════
ws_out = wb.create_sheet("Outputs")
ws_out.sheet_view.showGridLines = False
ws_out.column_dimensions["A"].width = 3
ws_out.column_dimensions["B"].width = 36
for col in ["C","D","E","F","G","H","I"]:
    ws_out.column_dimensions[col].width = 16

r = 1

# ── Summary of Results ────────────────────────────────────────────────────────
section_title(ws_out, r, 2, " DCF SCENARIO SUMMARY")
ws_out.merge_cells(f"B{r}:G{r}")
r += 1

header_row(ws_out, r,
    [(2,"Metric"),(3,"Bull Case"),(4,"Base Case"),(5,"Bear Case")],
    bg=C_DARK_BLUE)
r += 1

PRICE_NOW = 229.32
ANALYST   = 252.00

summary_rows = [
    ("WACC",                                    f"{WACC_BULL:.1%}", f"{WACC_BASE:.1%}", f"{WACC_BEAR:.1%}"),
    ("Terminal Growth Rate (g)",                f"{G_BULL:.1%}",    f"{G_BASE:.1%}",    f"{G_BEAR:.1%}"),
    ("Revenue CAGR 2025→2035",                  "5.9%",             "5.3%",             "3.5%"),
    ("2035E Revenue (USD B)",                   f"${bull_rev[-1]}", f"${base_rev[-1]}", f"${bear_rev[-1]}"),
    ("2035E FCF Margin",                        f"{bull_fcf_margins[-1]:.1%}", f"{base_fcf_margins[-1]:.1%}", f"{bear_fcf_margins[-1]:.1%}"),
    ("2035E Free Cash Flow (USD B)",            f"${bull_fcf_val[-1]}", f"${base_fcf_val[-1]}", f"${bear_fcf_val[-1]}"),
    ("PV of Explicit FCFs (10-yr, USD B)",      f"${bull_pv_fcf:.1f}", f"${base_pv_fcf:.1f}", f"${bear_pv_fcf:.1f}"),
    ("PV of Terminal Value (USD B)",            f"${bull_pv_tv:.1f}", f"${base_pv_tv:.1f}", f"${bear_pv_tv:.1f}"),
    ("Terminal Value % of EV",                  f"{bull_pv_tv/bull_ev:.1%}", f"{base_pv_tv/base_ev:.1%}", f"{bear_pv_tv/bear_ev:.1%}"),
    ("Enterprise Value (USD B)",                f"${bull_ev:.1f}",  f"${base_ev:.1f}",  f"${bear_ev:.1f}"),
    ("(−) Net Debt (USD B)",                    f"${NET_DEBT}",     f"${NET_DEBT}",     f"${NET_DEBT}"),
    ("Equity Value (USD B)",                    f"${bull_eq:.1f}",  f"${base_eq:.1f}",  f"${bear_eq:.1f}"),
    ("Equity Value per Share",                  f"${bull_ps:.1f}",  f"${base_ps:.1f}",  f"${bear_ps:.1f}"),
    (f"vs. Current Price (${PRICE_NOW})",       f"{(bull_ps/PRICE_NOW-1):+.1%}", f"{(base_ps/PRICE_NOW-1):+.1%}", f"{(bear_ps/PRICE_NOW-1):+.1%}"),
    (f"vs. Analyst Consensus (${ANALYST})",     f"{(bull_ps/ANALYST-1):+.1%}",   f"{(base_ps/ANALYST-1):+.1%}",   f"{(bear_ps/ANALYST-1):+.1%}"),
]
tones = [None,None,None,None,None,None,None,None,None,None,None,None,C_LIGHT_BLUE,None,None]
for i, (k, bv, bav, bev) in enumerate(summary_rows):
    label(ws_out, r, 2, k, bold=(k.startswith("Equity Value per")))
    for col, v in [(3, bv), (4, bav), (5, bev)]:
        c = ws_out.cell(row=r, column=col, value=v)
        c.font = font(bold=(k.startswith("Equity Value per")))
        c.alignment = align("right")
        if k.startswith("Equity Value per"):
            c.font = font(bold=True, size=11)
        bg_clr = C_LIGHT_GREEN if col==3 else (C_LIGHT_BLUE if col==4 else C_RED_LIGHT)
        c.fill = fill(bg_clr)
    r += 1

r += 2

# ── Sensitivity Table ─────────────────────────────────────────────────────────
section_title(ws_out, r, 2, " EQUITY VALUE PER SHARE — WACC × TERMINAL GROWTH SENSITIVITY  (USD)")
ws_out.merge_cells(f"B{r}:H{r}")
r += 1

ws_out.cell(row=r, column=2,
    value="Base-case FCF projections held constant. Net debt $32.9B deducted; 2.41B shares.")
ws_out.cell(row=r, column=2).font = font(italic=True, colour="595959")
r += 1

g_vals    = [0.020, 0.025, 0.030, 0.035, 0.040]
wacc_vals = [0.070, 0.075, 0.080, 0.085, 0.090, 0.095]

# PV of FCFs at each WACC (base FCFs)
pv_fcf_by_wacc = {w: pv_fcfs(base_fcf_val, w) for w in wacc_vals}
disc_10_by_wacc = {w: (1+w)**10 for w in wacc_vals}

# Header row
header_row(ws_out, r,
    [(2, "WACC \\ Terminal g")] + [(i+3, f"{g:.1%}") for i, g in enumerate(g_vals)],
    bg=C_DARK_BLUE)
r += 1

current_zone_highlighted = False
for wacc in wacc_vals:
    ws_out.cell(row=r, column=2, value=f"{wacc:.1%}").font = font(bold=True)
    ws_out.cell(row=r, column=2).fill = fill(C_GREY)
    for j, g in enumerate(g_vals):
        if wacc <= g:
            c = ws_out.cell(row=r, column=j+3, value="N/A (g≥WACC)")
            c.font = font(colour="FF0000")
            continue
        tv   = terminal_value(base_fcf_val[-1], wacc, g)
        ptv  = tv / disc_10_by_wacc[wacc]
        eq   = pv_fcf_by_wacc[wacc] + ptv - NET_DEBT
        ps_v = eq / SHARES
        c    = ws_out.cell(row=r, column=j+3, value=round(ps_v, 1))
        c.number_format = usd2_fmt()
        c.alignment = align("right")
        # Highlight near current price $229
        if abs(ps_v - PRICE_NOW) < 15:
            c.fill = fill(C_YELLOW)
            c.font = font(bold=True)
        # Highlight near analyst target $252
        elif abs(ps_v - ANALYST) < 15:
            c.fill = fill(C_LIGHT_GREEN)
        else:
            c.fill = fill(C_GREY if (wacc_vals.index(wacc)) % 2 == 0 else C_WHITE)
    r += 1

r += 1
ws_out.cell(row=r, column=2,
    value="Yellow cells: within $15 of current price ($229).  Green cells: within $15 of analyst target ($252).")
ws_out.cell(row=r, column=2).font = font(italic=True, colour="595959", size=9)
r += 2

# ── Verdict ───────────────────────────────────────────────────────────────────
section_title(ws_out, r, 2, " INVESTMENT VERDICT")
ws_out.merge_cells(f"B{r}:G{r}")
r += 1

verdict_rows = [
    ("VERDICT",                "HOLD — Do Not Add at Current Price $229",   C_ORANGE, C_YELLOW),
    ("Base DCF Fair Value",    f"${base_ps:.1f} per share",                 C_MID_BLUE, C_LIGHT_BLUE),
    ("Premium / Discount",     f"{(PRICE_NOW/base_ps-1):+.1%} vs base DCF (stock is slightly OVERVALUED)", C_MID_BLUE, C_LIGHT_BLUE),
    ("Buy-on-Dip Target",      "$195–$205 (near base DCF with margin of safety)", C_GREEN, C_LIGHT_GREEN),
    ("Trim / Take-Profit",     "$250–$260 (approaching analyst consensus)",  C_ORANGE, C_YELLOW),
    ("Dividend Yield Support", f"{3.05:.2f}% yield partially compensates for DCF premium", C_MID_BLUE, C_LIGHT_BLUE),
    ("Portfolio Role",         "Defensive anchor; beta 0.33 reduces overall portfolio volatility", C_MID_BLUE, C_LIGHT_BLUE),
    ("Your Position",          f"40 shares @ $168.90 avg cost → +{(229.32/168.90-1):.1%} unrealised gain. HOLD.", C_GREEN, C_LIGHT_GREEN),
]

for k, v, bg_k, bg_v in verdict_rows:
    ck = ws_out.cell(row=r, column=2, value=k)
    ck.font = font(bold=True, colour=C_WHITE)
    ck.fill = fill(bg_k)
    ck.alignment = align("left")
    cv = ws_out.cell(row=r, column=3, value=v)
    cv.font = font(bold=(k == "VERDICT"))
    cv.fill = fill(bg_v)
    cv.alignment = align("left", wrap=True)
    ws_out.merge_cells(f"C{r}:G{r}")
    ws_out.row_dimensions[r].height = 22
    r += 1

r += 2

# ── Analyst Consensus ─────────────────────────────────────────────────────────
section_title(ws_out, r, 2, " ANALYST CONSENSUS  (May 2026)")
ws_out.merge_cells(f"B{r}:G{r}")
r += 1

header_row(ws_out, r,
    [(2,"Firm"),(3,"Rating"),(4,"Price Target"),(5,"Implied Upside")],
    bg=C_DARK_BLUE)
r += 1

analyst_data = [
    ("Citigroup",       "Strong Buy", 285),
    ("HSBC",            "Strong Buy", 280),
    ("Morgan Stanley",  "Buy",        267),
    ("RBC Capital",     "Buy",        255),
    ("Barclays",        "Equal-Weight",255),
    ("Consensus (27a)", "Moderate Buy",252),
]
for firm, rating, tgt in analyst_data:
    ws_out.cell(row=r, column=2, value=firm)
    ws_out.cell(row=r, column=3, value=rating)
    c_tgt = ws_out.cell(row=r, column=4, value=tgt)
    c_tgt.number_format = usd2_fmt()
    c_tgt.alignment = align("right")
    c_up = ws_out.cell(row=r, column=5, value=tgt/PRICE_NOW-1)
    c_up.number_format = pct1_fmt()
    c_up.alignment = align("right")
    bg_row = C_LIGHT_GREEN if tgt >= PRICE_NOW else C_RED_LIGHT
    for col in [2,3,4,5]:
        ws_out.cell(row=r, column=col).fill = fill(bg_row)
    r += 1

r += 2
ws_out.cell(row=r, column=2,
    value="Sources: JNJ Q4 FY2025 earnings | Q1 2026 10-Q | MarketBeat | Yahoo Finance | Damodaran ERP | "
          "This model is for informational purposes only. Not investment advice.")
ws_out.cell(row=r, column=2).font = font(italic=True, colour="595959", size=9)
ws_out.merge_cells(f"B{r}:G{r}")

# ── Save ───────────────────────────────────────────────────────────────────────
out_path = r"c:\Users\KDWIVEDI\.cursor\DCF\JNJ DCF Valuation Model.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
